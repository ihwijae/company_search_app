import json
import math
import os
import re
from pathlib import Path

import xlwings as xw
from PySide6 import QtWidgets, QtCore
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "config.json"
_DB_CACHE = {"path": None, "mtime": None, "data": []}


def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_config(cfg):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def normalize_name(name: str) -> str:
    if not name:
        return ""
    name = str(name).strip().split("\n")[0]
    name = name.replace("(주)", "").replace("㈜", "").replace("주식회사", "")
    name = re.sub(r"\s*[0-9.,%].*$", "", name)
    return re.sub(r"\s+", " ", name).strip().lower()


def sanitize_company_name(name: str) -> str:
    if not name:
        return ""
    text = str(name).strip().split("\n")[0]
    text = text.replace("(주)", "").replace("㈜", "").replace("주식회사", "")
    return re.sub(r"\s+", " ", text).strip()


def extract_manager_name(notes: str):
    if not notes:
        return None
    text = re.sub(r"\s+", " ", str(notes)).strip()
    if not text:
        return None
    first_token = re.split(r"[ ,\/\|·\-]+", text)
    first_token = next((t for t in first_token if t), "")
    cleaned_first = re.sub(r"^[\[\(（【]([^\]\)）】]+)[\]\)】]?$", r"\1", first_token)
    if re.match(r"^[가-힣]{2,4}$", cleaned_first):
        return cleaned_first
    m = re.search(r"담당자?\s*[:：-]?\s*([가-힣]{2,4})", text)
    if m:
        return m.group(1)
    m = re.search(r"([가-힣]{2,4})\s*(과장|팀장|차장|대리|사원|부장|대표|실장|소장)", text)
    if m:
        return m.group(1)
    m = re.search(r"\b(?!확인서|등록증|증명서|평가|서류)([가-힣]{2,4})\b\s*(?:,|\/|\(|\d|$)", text)
    if m:
        return m.group(1)
    return None


def _to_number(val):
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return None
    s = str(val).replace(",", "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except Exception:
        return None


def load_db(db_path: Path):
    wb = load_workbook(db_path, data_only=False)
    data = []
    relative_offsets = {
        "대표자": 1,
        "사업자번호": 2,
        "지역": 3,
        "시평": 4,
        "3년 실적": 5,
        "5년 실적": 6,
        "부채비율": 7,
        "유동비율": 8,
        "영업기간": 9,
        "신용평가": 10,
        "여성기업": 11,
        "중소기업": 12,
        "일자리창출": 13,
        "품질평가": 14,
        "비고": 15,
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        merged_value = {}
        for merged in ws.merged_cells.ranges:
            tl = ws.cell(merged.min_row, merged.min_col).value
            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    merged_value[(r, c)] = tl

        def get_value(r, c):
            v = ws.cell(r, c).value
            if v is None:
                return merged_value.get((r, c))
            return v

        header_positions = []
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = get_value(r, c)
                if cell is None:
                    continue
                val = str(cell)
                if "회사명" in val:
                    header_positions.append((r, c))

        if not header_positions:
            continue

        seen_keys = set()
        for header_row, header_col in header_positions:
            for col in range(header_col + 1, max_col + 1):
                raw_name = get_value(header_row, col)
                if raw_name is None:
                    continue
                raw_name = str(raw_name).strip()
                if not raw_name:
                    continue
                name = raw_name.split("\n")[0].strip()
                if not name:
                    continue
                dedup_key = (sheet_name, header_row, col, name)
                if dedup_key in seen_keys:
                    continue
                seen_keys.add(dedup_key)
            entry = {
                "name": name,
                "norm": normalize_name(name),
                "region": sheet_name.strip(),
                "bizNo": "",
                "debtRatio": None,
                "currentRatio": None,
                "perf5y": None,
                "creditGrade": "",
                "sipyung": None,
                "notes": "",
                "managerName": "",
            }
            for key, offset in relative_offsets.items():
                r = header_row + offset
                if r > max_row:
                    continue
                val = get_value(r, col)
                if key in {"부채비율", "유동비율"} and isinstance(val, (int, float)):
                    val = val * 100
                if key == "사업자번호":
                    entry["bizNo"] = "" if val is None else str(val).strip()
                elif key == "부채비율":
                    entry["debtRatio"] = _to_number(val)
                elif key == "유동비율":
                    entry["currentRatio"] = _to_number(val)
                elif key == "시평":
                    entry["sipyung"] = _to_number(val)
                elif key == "5년 실적":
                    entry["perf5y"] = _to_number(val)
                elif key == "신용평가":
                    entry["creditGrade"] = "" if val is None else str(val).strip()
                elif key == "비고":
                    entry["notes"] = "" if val is None else str(val).strip()
            entry["managerName"] = extract_manager_name(entry.get("notes", ""))
            data.append(entry)
    return data


def load_db_cached(db_path: Path, force=False):
    mtime = db_path.stat().st_mtime if db_path.exists() else None
    if (
        not force
        and _DB_CACHE["path"] == str(db_path)
        and _DB_CACHE["mtime"] == mtime
    ):
        return _DB_CACHE["data"]
    data = load_db(db_path)
    _DB_CACHE["path"] = str(db_path)
    _DB_CACHE["mtime"] = mtime
    _DB_CACHE["data"] = data
    return data


def _truncate(value, digits=2):
    factor = 10 ** digits
    return int(value * factor) / factor


def score_debt_mois_under30(ratio: float):
    if ratio < 0.5:
        return 8.0
    if ratio < 0.75:
        return 7.2
    if ratio < 1.0:
        return 6.4
    if ratio < 1.25:
        return 5.6
    return 4.8


def score_current_mois_under30(ratio: float):
    if ratio >= 1.5:
        return 7.0
    if ratio >= 1.2:
        return 6.3
    if ratio >= 1.0:
        return 5.6
    if ratio >= 0.7:
        return 4.9
    return 4.2


def score_credit_mois_under30(grade: str):
    g = str(grade).strip().upper()
    if g in {"AAA", "AA+", "AA0", "AA-", "A+", "A0", "A-", "BBB+", "BBB0", "BBB-", "BB+", "BB0"}:
        return 15
    if g == "BB-":
        return 14
    if g in {"B+", "B0", "B-"}:
        return 13
    if g in {"CCC+", "CCC0", "CCC-", "CC", "C", "D"}:
        return 10
    return None


def compute_management_mois_under30(row, file_type, industry_avg):
    debt = row.get("debtRatio")
    current = row.get("currentRatio")
    debt_avg = industry_avg[file_type]["debtRatio"]
    current_avg = industry_avg[file_type]["currentRatio"]

    composite = 0.0
    if debt is not None and debt_avg:
        composite += score_debt_mois_under30(debt / debt_avg)
    if current is not None and current_avg:
        composite += score_current_mois_under30(current / current_avg)

    credit = score_credit_mois_under30(row.get("creditGrade", ""))
    best = max(composite, credit or 0)
    best = min(15.0, max(0.0, best))
    return _truncate(best, 2)


def write_to_active_cell(value):
    book = xw.Book.caller()
    rng = book.app.selection
    rng.value = value


def apply_mois_under30(name_value, row_data, file_type):
    cfg = load_config()
    industry_avg = cfg["industryAverages"]

    book = xw.Book.caller()
    sht = book.sheets.active

    # Determine slot based on active cell column
    settings = cfg["mois_under30"]
    name_cols = settings["nameCols"]
    mgmt_cols = settings["managementCols"]
    perf_cols = settings["performanceCols"]
    sipyung_cols = settings.get("sipyungCols", [])

    active = book.app.selection
    col_letter = re.sub(r"\d", "", active.address.split("$")[-1])
    row_num = active.row

    if col_letter not in name_cols:
        return
    idx = name_cols.index(col_letter)

    mgmt = compute_management_mois_under30(row_data, file_type, industry_avg)
    if mgmt is not None:
        sht.range(f"{mgmt_cols[idx]}{row_num}").value = mgmt
    perf = row_data.get("perf5y")
    if perf is not None:
        sht.range(f"{perf_cols[idx]}{row_num}").value = perf
    sipyung = row_data.get("sipyung")
    if sipyung is not None and idx < len(sipyung_cols):
        sht.range(f"{sipyung_cols[idx]}{row_num}").value = sipyung


def open_modal():
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([])

    cfg = load_config()
    db_path = Path(cfg.get("dbPath", ""))
    if not db_path.is_absolute():
        db_path = (BASE_DIR / db_path).resolve()

    if not db_path.exists():
        QtWidgets.QMessageBox.warning(None, "DB 경로", "DB 파일 경로를 설정하세요.")
        picked, _ = QtWidgets.QFileDialog.getOpenFileName(
            None,
            "업체 DB 선택",
            str(BASE_DIR),
            "Excel Files (*.xlsx)",
        )
        if not picked:
            return
        cfg["dbPath"] = os.path.relpath(picked, BASE_DIR)
        save_config(cfg)
        db_path = Path(picked)

    data = load_db_cached(db_path)

    dialog = QtWidgets.QDialog()
    dialog.setWindowTitle("업체 검색")
    dialog.resize(720, 520)

    layout = QtWidgets.QVBoxLayout(dialog)
    status_label = QtWidgets.QLabel(f"DB: {db_path} (로드 {len(data)}건)")
    layout.addWidget(status_label)

    form = QtWidgets.QHBoxLayout()
    industry_box = QtWidgets.QComboBox()
    industry_box.addItems(["전기", "통신", "소방"])
    form.addWidget(QtWidgets.QLabel("공종"))
    form.addWidget(industry_box)

    query_input = QtWidgets.QLineEdit()
    form.addWidget(QtWidgets.QLabel("업체명"))
    form.addWidget(query_input)

    search_btn = QtWidgets.QPushButton("검색")
    form.addWidget(search_btn)

    config_btn = QtWidgets.QPushButton("DB 경로 설정")
    form.addWidget(config_btn)

    verify_btn = QtWidgets.QPushButton("DB 경로 확인")
    form.addWidget(verify_btn)

    layout.addLayout(form)

    table = QtWidgets.QTableWidget(0, 3)
    table.setHorizontalHeaderLabels(["업체명", "지역", "사업자번호"])
    table.horizontalHeader().setStretchLastSection(True)
    layout.addWidget(table)

    def do_search():
        q = normalize_name(query_input.text())
        table.setRowCount(0)
        if not q:
            return
        for row in data:
            if q in row["norm"]:
                r = table.rowCount()
                table.insertRow(r)
                table.setItem(r, 0, QtWidgets.QTableWidgetItem(row["name"]))
                table.setItem(r, 1, QtWidgets.QTableWidgetItem(row["region"]))
                table.setItem(r, 2, QtWidgets.QTableWidgetItem(row.get("bizNo", "")))

    def apply_selected():
        selected = table.currentRow()
        if selected < 0:
            return
        name_val = table.item(selected, 0).text()
        region_val = table.item(selected, 1).text()
        biz_val = table.item(selected, 2).text()
        row_data = next((r for r in data if r["name"] == name_val and r["region"] == region_val), None)
        if not row_data:
            return
        clean_name = sanitize_company_name(name_val) or name_val
        manager_name = row_data.get("managerName", "")
        display_name = f"{clean_name}\n{manager_name}".strip() if manager_name else clean_name
        write_to_active_cell(display_name)

        file_type = {
            "전기": "eung",
            "통신": "tongsin",
            "소방": "sobang",
        }[industry_box.currentText()]
        apply_mois_under30(name_val, row_data, file_type)
        dialog.accept()

    def set_db_path():
        nonlocal data, db_path
        path, _ = QtWidgets.QFileDialog.getOpenFileName(dialog, "업체 DB 선택", str(BASE_DIR), "Excel Files (*.xlsx)")
        if not path:
            return
        cfg["dbPath"] = os.path.relpath(path, BASE_DIR)
        save_config(cfg)
        db_path = Path(path)
        data = load_db_cached(db_path, force=True)
        status_label.setText(f"DB: {db_path} (로드 {len(data)}건)")
        QtWidgets.QMessageBox.information(dialog, "DB 경로", f"설정됨:\n{db_path}\n로드 {len(data)}건")

    def verify_db_path():
        exists = db_path.exists()
        mtime = db_path.stat().st_mtime if exists else None
        msg = f"경로: {db_path}\n존재: {'예' if exists else '아니오'}\n로드 {len(data)}건"
        if mtime:
            msg += f"\n수정시간: {QtCore.QDateTime.fromSecsSinceEpoch(int(mtime)).toString('yyyy-MM-dd HH:mm:ss')}"
        QtWidgets.QMessageBox.information(dialog, "DB 경로 확인", msg)

    search_btn.clicked.connect(do_search)
    query_input.returnPressed.connect(do_search)
    config_btn.clicked.connect(set_db_path)
    verify_btn.clicked.connect(verify_db_path)
    table.itemDoubleClicked.connect(lambda _: apply_selected())

    btns = QtWidgets.QHBoxLayout()
    apply_btn = QtWidgets.QPushButton("선택")
    close_btn = QtWidgets.QPushButton("닫기")
    btns.addStretch(1)
    btns.addWidget(apply_btn)
    btns.addWidget(close_btn)
    layout.addLayout(btns)

    apply_btn.clicked.connect(apply_selected)
    close_btn.clicked.connect(dialog.reject)

    dialog.exec()
